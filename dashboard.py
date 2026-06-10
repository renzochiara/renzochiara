import os
import streamlit as st
from mistralai import Mistral  # Updated import
import pandas as pd
from openpyxl import load_workbook
import pdfplumber
from dotenv import load_dotenv

# Load environment variables
load_dotenv()
MISTRAL_API_KEY = os.getenv("MISTRAL_API_KEY")

# Initialize Mistral client
client = Mistral(api_key=MISTRAL_API_KEY)

# --- Streamlit App ---
st.set_page_config(page_title="MyVita Budget Dashboard", layout="wide")

# Basic email/password protection
def check_password():
    return st.secrets["EMAIL"] == st.session_state.get("email", "") and st.secrets["PASSWORD"] == st.session_state.get("password", "")

if "logged_in" not in st.session_state:
    st.session_state.logged_in = False

if not st.session_state.logged_in:
    st.subheader("Login")
    st.session_state.email = st.text_input("Email", key="login_email")
    st.session_state.password = st.text_input("Password", type="password", key="login_password")
    if st.button("Login"):
        if check_password():
            st.session_state.logged_in = True
            st.rerun()
        else:
            st.error("Invalid email or password")
else:
    st.title("🚀 MyVita Budget Automation Dashboard")
    st.markdown("Upload files, chat with Mistral AI, and run workflows for CMO142 and other parkings.")

    # --- Sidebar for Navigation ---
    st.sidebar.title("Navigation")
    page = st.sidebar.radio("Go to", ["Chat with AI", "Upload & Process Files", "Run Workflow", "Logout"])

    if page == "Logout":
        st.session_state.logged_in = False
        st.rerun()

    # --- Chat with Mistral AI ---
    if page == "Chat with AI":
        st.subheader("💬 Chat with Mistral AI")
        user_input = st.text_input("Ask me anything or request a workflow:")
        if user_input:
            try:
                response = client.chat(
                    model="mistral-medium",
                    messages=[{"role": "user", "content": user_input}]
                )
                st.write("**Mistral AI:**")
                st.write(response.choices[0].message.content)
            except Exception as e:
                st.error(f"Error: {e}")

    # --- Upload & Process Files ---
    elif page == "Upload & Process Files":
        st.subheader("📁 Upload Files")
        uploaded_excel = st.file_uploader("Upload Excel Template", type=["xlsx"])
        uploaded_pdf = st.file_uploader("Upload PDF Report", type=["pdf"])

        if uploaded_excel and uploaded_pdf:
            st.success("Files uploaded successfully!")
            st.write("**Excel File:**", uploaded_excel.name)
            st.write("**PDF File:**", uploaded_pdf.name)

            # Preview PDF data
            try:
                with pdfplumber.open(uploaded_pdf) as pdf:
                    pnl_table = pdf.pages[0].extract_table()
                st.write("**PDF Preview (First Table):**")
                st.dataframe(pd.DataFrame(pnl_table[1:], columns=pnl_table[0]))
            except Exception as e:
                st.error(f"Error reading PDF: {e}")

    # --- Run Workflow ---
    elif page == "Run Workflow":
        st.subheader("⚙️ Run Budget Workflow")
        st.write("Select a parking lot and run the workflow to generate the budget.")

        # Parking lot selection
        parking_lots = ["CMO142 (LUNA)", "CMO143", "CMO144"]
        selected_parking = st.selectbox("Select Parking Lot", parking_lots)

        # Parking type selection
        parking_type = st.radio("Parking Type", ["SC (School)", "RG (Regular/Tourism)"])

        # Supervisor hours input
        supervisor_hours = st.number_input("Supervisor Hours/Day", min_value=0, max_value=24, value=1)

        if st.button(f"Generate Budget for {selected_parking}"):
            if uploaded_excel and uploaded_pdf:
                st.write("Processing files...")

                try:
                    # Extract data from PDF
                    with pdfplumber.open(uploaded_pdf) as pdf:
                        pnl_table = pdf.pages[0].extract_table()
                    df_pnl = pd.DataFrame(pnl_table[1:], columns=pnl_table[0])

                    # Load Excel template
                    wb = load_workbook(uploaded_excel)
                    ws = wb["1-Fiche Stationnement"]

                    # Fill K17-K26 (case-insensitive and flexible column matching)
                    for row in df_pnl.itertuples(index=False):
                        if "transient" in str(row[0]).lower() and "revenue" in str(row[0]).lower():
                            ws["K17"] = float(row[1])
                        elif "monthly" in str(row[0]).lower() and "revenue" in str(row[0]).lower():
                            ws["K18"] = float(row[1])
                        elif "total" in str(row[0]).lower() and "revenue" in str(row[0]).lower():
                            ws["K26"] = float(row[1])

                    # Apply seasonal multipliers (based on parking_type)
                    if parking_type == "SC (School)":
                        multipliers = {
                            "Jan": 1.2, "Feb": 1.2, "Mar": 1.2, "Apr": 1.1,
                            "May": 0.8, "Jun": 0.8, "Jul": 0.8, "Aug": 0.8,
                            "Sep": 1.1, "Oct": 1.1, "Nov": 1.2, "Dec": 1.2
                        }
                    else:  # RG (Regular/Tourism)
                        multipliers = {
                            "Jan": 0.8, "Feb": 0.8, "Mar": 0.9, "Apr": 0.9,
                            "May": 1.3, "Jun": 1.3, "Jul": 1.3, "Aug": 1.2,
                            "Sep": 1.0, "Oct": 1.0, "Nov": 0.8, "Dec": 0.8
                        }

                    # Apply supervisor costs
                    hourly_rate = 25.0
                    monthly_cost = hourly_rate * supervisor_hours * 30
                    if "4-Feuille calcul-Salaire" in wb.sheetnames:
                        ws_salary = wb["4-Feuille calcul-Salaire"]
                        ws_salary["D10"] = monthly_cost

                    # Save the updated file
                    output_path = f"{selected_parking.replace(' ', '_')}_budget.xlsx"
                    wb.save(output_path)

                    # Provide download link
                    with open(output_path, "rb") as f:
                        st.download_button(
                            label="Download Budget File",
                            data=f,
                            file_name=output_path,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                    st.success(f"Budget generated for {selected_parking}!")
                except Exception as e:
                    st.error(f"Error processing files: {e}")
            else:
                st.error("Please upload both Excel and PDF files.")
        
